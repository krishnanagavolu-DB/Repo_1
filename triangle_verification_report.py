import argparse
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PWC_FILE = "pwc_report.csv"
VHQ_FILE = "vhq_report.csv"
ITRAC_FILE = "itrac_report.csv"
DEFAULT_OUTPUT_FILE = "mismatch_alert_report.xlsx"

REQUIRED_COLUMNS = {
    PWC_FILE: ["NewCoID", "Merchant_ID", "Terminal_ID", "Serial_Number", "Status"],
    VHQ_FILE: ["NewCoID", "Serial_Number", "Hierarchy", "Status", "Last_Heartbeat"],
    ITRAC_FILE: ["NewCoID", "Shipped_Serial_Number", "Shipping_Status"],
}

DETAIL_COLUMNS = [
    "Severity",
    "NewCoID",
    "Serial_Number",
    "Rule_Violated",
    "Specific_Issue",
    "Source_System",
    "PWC_Status",
    "VHQ_Status",
    "iTrac_Shipping_Status",
]


def clean_value(value: object) -> str:
    """Convert empty cells and numbers into clean text for matching."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_status(value: object) -> str:
    """Normalize small formatting differences like Green / Active."""
    status = clean_value(value).lower()
    status = re.sub(r"\s*/\s*", "/", status)
    status = re.sub(r"\s+", " ", status)
    return status


def load_report(input_dir: Path, filename: str) -> pd.DataFrame:
    path = input_dir / filename
    if not path.exists():
        raise FileNotFoundError(
            f"Missing {filename}. Place it in {input_dir.resolve()} and try again."
        )

    report = pd.read_csv(path, dtype=str, keep_default_na=False)
    report.columns = [clean_value(column) for column in report.columns]

    missing_columns = [
        column for column in REQUIRED_COLUMNS[filename] if column not in report.columns
    ]
    if missing_columns:
        missing_list = ", ".join(missing_columns)
        raise ValueError(f"{filename} is missing required column(s): {missing_list}")

    return report


def prepare_report(report: pd.DataFrame, serial_column: str) -> pd.DataFrame:
    prepared = report.copy()
    prepared["NewCoID"] = prepared["NewCoID"].map(clean_value)
    prepared[serial_column] = prepared[serial_column].map(clean_value)
    if "Status" in prepared.columns:
        prepared["Status"] = prepared["Status"].map(clean_value)
    if "Shipping_Status" in prepared.columns:
        prepared["Shipping_Status"] = prepared["Shipping_Status"].map(clean_value)
    return prepared


def serials_by_newco(report: pd.DataFrame, serial_column: str) -> dict[str, set[str]]:
    grouped: dict[str, set[str]] = {}
    for newco_id, serial in zip(report["NewCoID"], report[serial_column]):
        if not newco_id or not serial:
            continue
        grouped.setdefault(newco_id, set()).add(serial)
    return grouped


def first_status_by_device(
    report: pd.DataFrame, serial_column: str, status_column: str
) -> dict[tuple[str, str], str]:
    statuses: dict[tuple[str, str], str] = {}
    for _, row in report.iterrows():
        newco_id = clean_value(row["NewCoID"])
        serial = clean_value(row[serial_column])
        status = clean_value(row[status_column])
        if not newco_id or not serial:
            continue
        statuses.setdefault((newco_id, serial), status)
    return statuses


def add_issue(
    issues: list[dict[str, str]],
    *,
    severity: str,
    newco_id: str,
    serial: str,
    rule: str,
    issue: str,
    source_system: str,
    pwc_status: str = "",
    vhq_status: str = "",
    itrac_shipping_status: str = "",
) -> None:
    issues.append(
        {
            "Severity": severity,
            "NewCoID": newco_id or "(blank)",
            "Serial_Number": serial or "(blank)",
            "Rule_Violated": rule,
            "Specific_Issue": issue,
            "Source_System": source_system,
            "PWC_Status": pwc_status,
            "VHQ_Status": vhq_status,
            "iTrac_Shipping_Status": itrac_shipping_status,
        }
    )


def apply_rule_1(
    issues: list[dict[str, str]],
    pwc_serials: dict[str, set[str]],
    vhq_serials: dict[str, set[str]],
    pwc_statuses: dict[tuple[str, str], str],
    vhq_statuses: dict[tuple[str, str], str],
) -> None:
    all_newco_ids = sorted(set(pwc_serials) | set(vhq_serials))
    for newco_id in all_newco_ids:
        pwc_only = sorted(pwc_serials.get(newco_id, set()) - vhq_serials.get(newco_id, set()))
        vhq_only = sorted(vhq_serials.get(newco_id, set()) - pwc_serials.get(newco_id, set()))

        for serial in pwc_only:
            add_issue(
                issues,
                severity="High",
                newco_id=newco_id,
                serial=serial,
                rule="Rule 1 - PWC vs VHQ Alignment",
                issue="Serial number exists in PWC but is missing from VHQ for this NewCoID.",
                source_system="PWC",
                pwc_status=pwc_statuses.get((newco_id, serial), ""),
                vhq_status=vhq_statuses.get((newco_id, serial), ""),
            )

        for serial in vhq_only:
            add_issue(
                issues,
                severity="High",
                newco_id=newco_id,
                serial=serial,
                rule="Rule 1 - PWC vs VHQ Alignment",
                issue="Serial number exists in VHQ but is missing from PWC for this NewCoID.",
                source_system="VHQ",
                pwc_status=pwc_statuses.get((newco_id, serial), ""),
                vhq_status=vhq_statuses.get((newco_id, serial), ""),
            )


def apply_rule_2(
    issues: list[dict[str, str]],
    itrac_report: pd.DataFrame,
    pwc_serials: dict[str, set[str]],
    vhq_serials: dict[str, set[str]],
    pwc_statuses: dict[tuple[str, str], str],
    vhq_statuses: dict[tuple[str, str], str],
) -> None:
    rule_name = "Rule 2 - Triangle Check"
    itrac_devices = itrac_report.drop_duplicates(
        subset=["NewCoID", "Shipped_Serial_Number", "Shipping_Status"]
    )

    for _, row in itrac_devices.iterrows():
        newco_id = clean_value(row["NewCoID"])
        serial = clean_value(row["Shipped_Serial_Number"])
        shipping_status = clean_value(row["Shipping_Status"])

        if not serial:
            add_issue(
                issues,
                severity="Critical",
                newco_id=newco_id,
                serial=serial,
                rule=rule_name,
                issue="iTrac has a shipped-device row with a blank serial number.",
                source_system="iTrac",
                itrac_shipping_status=shipping_status,
            )
            continue

        missing_systems = []
        if serial not in pwc_serials.get(newco_id, set()):
            missing_systems.append("PWC")
        if serial not in vhq_serials.get(newco_id, set()):
            missing_systems.append("VHQ")

        if missing_systems:
            missing_text = " and ".join(missing_systems)
            add_issue(
                issues,
                severity="Critical",
                newco_id=newco_id,
                serial=serial,
                rule=rule_name,
                issue=(
                    "Physical serial number is shipped in iTrac but is missing from "
                    f"{missing_text} digital mapping for this NewCoID."
                ),
                source_system="iTrac",
                pwc_status=pwc_statuses.get((newco_id, serial), ""),
                vhq_status=vhq_statuses.get((newco_id, serial), ""),
                itrac_shipping_status=shipping_status,
            )


def apply_rule_3(
    issues: list[dict[str, str]],
    pwc_report: pd.DataFrame,
    vhq_report: pd.DataFrame,
    vhq_statuses: dict[tuple[str, str], str],
) -> None:
    rule_name = "Rule 3 - Status Check"

    for _, row in pwc_report.drop_duplicates(subset=["NewCoID", "Serial_Number", "Status"]).iterrows():
        newco_id = clean_value(row["NewCoID"])
        serial = clean_value(row["Serial_Number"])
        status = clean_value(row["Status"])

        if clean_status(status) != "green/active":
            add_issue(
                issues,
                severity="High",
                newco_id=newco_id,
                serial=serial,
                rule=rule_name,
                issue='PWC status is not "Green/Active".',
                source_system="PWC",
                pwc_status=status,
                vhq_status=vhq_statuses.get((newco_id, serial), ""),
            )

    bad_vhq_statuses = {"inactive", "pending registration"}
    for _, row in vhq_report.drop_duplicates(subset=["NewCoID", "Serial_Number", "Status"]).iterrows():
        newco_id = clean_value(row["NewCoID"])
        serial = clean_value(row["Serial_Number"])
        status = clean_value(row["Status"])

        if clean_status(status) in bad_vhq_statuses:
            add_issue(
                issues,
                severity="High",
                newco_id=newco_id,
                serial=serial,
                rule=rule_name,
                issue='VHQ status is "Inactive" or "Pending Registration".',
                source_system="VHQ",
                vhq_status=status,
            )


def build_summary(details: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    total_risks = len(details)
    critical_risks = int((details["Severity"] == "Critical").sum()) if total_risks else 0
    impacted_newco_ids = details["NewCoID"].nunique() if total_risks else 0

    overview = pd.DataFrame(
        [
            {"Metric": "Total risks flagged", "Value": total_risks},
            {"Metric": "Critical risks", "Value": critical_risks},
            {"Metric": "Impacted NewCoIDs", "Value": impacted_newco_ids},
            {
                "Metric": "Report meaning",
                "Value": "Rows in Details show the exact NewCoID, serial number, and rule violated.",
            },
        ]
    )

    if total_risks:
        by_rule = (
            details.groupby("Rule_Violated")
            .size()
            .reset_index(name="Issue_Count")
            .sort_values("Issue_Count", ascending=False)
        )
        by_severity = (
            details.groupby("Severity")
            .size()
            .reset_index(name="Issue_Count")
            .sort_values("Issue_Count", ascending=False)
        )
    else:
        by_rule = pd.DataFrame(columns=["Rule_Violated", "Issue_Count"])
        by_severity = pd.DataFrame(columns=["Severity", "Issue_Count"])

    return overview, by_rule, by_severity


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def style_header_row(worksheet, row_number: int) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in worksheet[row_number]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def style_workbook(output_path: Path, details: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)

    summary_sheet = workbook["Summary"]
    style_header_row(summary_sheet, 1)
    for row_number in range(1, summary_sheet.max_row + 1):
        summary_sheet.row_dimensions[row_number].height = 20
    autosize_columns(summary_sheet)

    details_sheet = workbook["Details"]
    style_header_row(details_sheet, 1)
    details_sheet.freeze_panes = "A2"
    details_sheet.auto_filter.ref = details_sheet.dimensions

    severity_fills = {
        "Critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "High": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }
    severity_column = DETAIL_COLUMNS.index("Severity") + 1
    for row_number in range(2, details_sheet.max_row + 1):
        severity = details_sheet.cell(row=row_number, column=severity_column).value
        fill = severity_fills.get(severity)
        if fill:
            for column_number in range(1, details_sheet.max_column + 1):
                details_sheet.cell(row=row_number, column=column_number).fill = fill

    for row in details_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_columns(details_sheet)
    if details.empty:
        details_sheet["A2"] = "No mismatches found."

    workbook.save(output_path)


def write_report(issues: Iterable[dict[str, str]], output_path: Path) -> None:
    details = pd.DataFrame(list(issues), columns=DETAIL_COLUMNS)
    overview, by_rule, by_severity = build_summary(details)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Summary", index=False, startrow=0)

        rule_start_row = len(overview) + 3
        pd.DataFrame([{"Breakdown": "Risks by rule"}]).to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            header=False,
            startrow=rule_start_row - 1,
        )
        by_rule.to_excel(writer, sheet_name="Summary", index=False, startrow=rule_start_row)

        severity_start_row = rule_start_row + len(by_rule) + 4
        pd.DataFrame([{"Breakdown": "Risks by severity"}]).to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            header=False,
            startrow=severity_start_row - 1,
        )
        by_severity.to_excel(
            writer, sheet_name="Summary", index=False, startrow=severity_start_row
        )

        details.to_excel(writer, sheet_name="Details", index=False)

    style_workbook(output_path, details)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run Triangle Verification for Verifone serial numbers."
    )
    parser.add_argument(
        "--input-dir",
        default=".",
        help="Folder containing pwc_report.csv, vhq_report.csv, and itrac_report.csv.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output file name. Default: {DEFAULT_OUTPUT_FILE}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output)

    pwc_report = prepare_report(load_report(input_dir, PWC_FILE), "Serial_Number")
    vhq_report = prepare_report(load_report(input_dir, VHQ_FILE), "Serial_Number")
    itrac_report = prepare_report(load_report(input_dir, ITRAC_FILE), "Shipped_Serial_Number")

    pwc_serials = serials_by_newco(pwc_report, "Serial_Number")
    vhq_serials = serials_by_newco(vhq_report, "Serial_Number")
    pwc_statuses = first_status_by_device(pwc_report, "Serial_Number", "Status")
    vhq_statuses = first_status_by_device(vhq_report, "Serial_Number", "Status")

    issues: list[dict[str, str]] = []
    apply_rule_1(issues, pwc_serials, vhq_serials, pwc_statuses, vhq_statuses)
    apply_rule_2(
        issues,
        itrac_report,
        pwc_serials,
        vhq_serials,
        pwc_statuses,
        vhq_statuses,
    )
    apply_rule_3(issues, pwc_report, vhq_report, vhq_statuses)

    write_report(issues, output_path)
    print(f"Created {output_path}")
    print(f"Total risks flagged: {len(issues)}")


if __name__ == "__main__":
    main()
